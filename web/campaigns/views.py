from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages as django_messages
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db import transaction, IntegrityError
from .models import Campaign, CampaignLog, Contact
import sys
import os
import threading
import openpyxl
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from runner.campaign_runner import run_campaign

COLUMN_MAP = {
    'Business Name': 'business_name',
    'Owner Name': 'owner_name',
    'Phone': 'phone',
    'City': 'city',
    'Category': 'category',
    'Email': 'email',
}

REQUIRED_COLUMNS = {'Business Name', 'Phone'}


def dashboard(request):
    campaigns = Campaign.objects.all()
    logs = {log.campaign_id: log for log in CampaignLog.objects.order_by('-completed_at')}
    for campaign in campaigns:
        campaign.log = logs.get(campaign.id)
    has_running_campaigns = campaigns.filter(status='running').exists()
    return render(request, 'campaigns/dashboard.html', {
        'campaigns': campaigns,
        'has_running_campaigns': has_running_campaigns,
    })


def new_campaign(request):
    cities = Contact.objects.values_list('city', flat=True).distinct().exclude(city=None)
    categories = Contact.objects.values_list('category', flat=True).distinct().exclude(category=None)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        template = request.POST.get('template', '').strip()
        city = request.POST.get('city', '').strip()
        category = request.POST.get('category', '').strip()

        if not name or not template:
            django_messages.error(request, 'Campaign name and message template are required.')
            return render(request, 'campaigns/new_campaign.html', {
                'cities': cities,
                'categories': categories,
            })

        filters = {}
        if city:
            filters['city'] = city
        if category:
            filters['category'] = category

        try:
            with transaction.atomic():
                campaign = Campaign.objects.create(
                    name=name,
                    template=template,
                    segment_filter=filters if filters else {},
                    status='running',  # reserved atomically — DB constraint blocks a second concurrent 'running' row
                )

            thread = threading.Thread(
                target=run_campaign,
                kwargs={
                    'campaign_id': str(campaign.id),
                    'template': template,
                    'filters': filters if filters else None,
                },
                daemon=True
            )
            thread.start()

            django_messages.success(request, f'Campaign "{name}" launched. Check dashboard for status.')
            return redirect('dashboard')

        except IntegrityError:
            django_messages.error(
                request,
                'A campaign is already running. Wait for it to finish before launching another.'
            )
            return render(request, 'campaigns/new_campaign.html', {
                'cities': cities,
                'categories': categories,
            })

        except Exception as e:
            django_messages.error(request, f'Failed to launch campaign: {str(e)}')
            return render(request, 'campaigns/new_campaign.html', {
                'cities': cities,
                'categories': categories,
            })

    return render(request, 'campaigns/new_campaign.html', {
        'cities': cities,
        'categories': categories,
    })


def campaign_detail(request, campaign_id):
    campaign = get_object_or_404(Campaign, id=campaign_id)
    log = CampaignLog.objects.filter(campaign=campaign).order_by('-completed_at').first()
    return render(request, 'campaigns/campaign_detail.html', {
        'campaign': campaign,
        'log': log,
    })


def download_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    headers = ['Business Name', 'Owner Name', 'Phone', 'City', 'Category', 'Email']

    # Style header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = 20

    # Example row so marketing team understands the format
    ws.append([
        'Sharma Traders',
        'Ram Sharma',
        '+9779812345678',
        'Pokhara',
        'retail',
        'ram@example.com'
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contacts_template.xlsx"'
    wb.save(response)
    return response


def import_contacts(request):
    if request.method == 'POST':
        file = request.FILES.get('file')

        if not file:
            django_messages.error(request, 'No file selected.')
            return render(request, 'campaigns/import_contacts.html')

        if not file.name.endswith('.xlsx'):
            django_messages.error(request, 'Invalid file type. Please upload a .xlsx file.')
            return render(request, 'campaigns/import_contacts.html')

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # Read headers from first row
            headers = [cell.value for cell in ws[1]]

            # Check required columns exist
            missing = REQUIRED_COLUMNS - set(headers)
            if missing:
                django_messages.error(
                    request,
                    f'Missing required columns: {", ".join(missing)}. '
                    f'Please download the template and use it.'
                )
                return render(request, 'campaigns/import_contacts.html')

            # Check for unrecognized columns
            unrecognized = [h for h in headers if h and h not in COLUMN_MAP]
            if unrecognized:
                django_messages.error(
                    request,
                    f'Unrecognized columns: {", ".join(unrecognized)}. '
                    f'Please download the template and use it.'
                )
                return render(request, 'campaigns/import_contacts.html')

            imported = 0
            skipped = 0
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip completely empty rows
                if not any(row):
                    continue

                row_data = dict(zip(headers, row))

                # Map to DB field names
                contact_data = {}
                for excel_col, db_field in COLUMN_MAP.items():
                    value = row_data.get(excel_col)
                    if value is not None:
                        contact_data[db_field] = str(value).strip()

                # Validate required fields
                if not contact_data.get('business_name'):
                    errors.append(f'Row {row_num}: Missing Business Name — skipped.')
                    skipped += 1
                    continue

                if not contact_data.get('phone'):
                    errors.append(f'Row {row_num}: Missing Phone — skipped.')
                    skipped += 1
                    continue

                # Normalize city casing
                if contact_data.get('city'):
                    contact_data['city'] = contact_data['city'].title()

                # Insert or skip duplicate phone
                if Contact.objects.filter(phone=contact_data['phone']).exists():
                    errors.append(f'Row {row_num}: {contact_data["phone"]} already exists — skipped.')
                    skipped += 1
                    continue

                Contact.objects.create(**contact_data)
                imported += 1

            # Build result message
            if imported > 0:
                django_messages.success(
                    request,
                    f'Import complete. {imported} contacts added, {skipped} skipped.'
                )
            else:
                django_messages.error(
                    request,
                    f'No contacts imported. {skipped} rows skipped.'
                )

            if errors:
                for error in errors[:5]:  # Show max 5 errors to avoid flooding UI
                    django_messages.warning(request, error)
                if len(errors) > 5:
                    django_messages.warning(request, f'...and {len(errors) - 5} more issues.')

            return redirect('import_contacts')

        except Exception as e:
            django_messages.error(request, f'Failed to read file: {str(e)}')
            return render(request, 'campaigns/import_contacts.html')

    return render(request, 'campaigns/import_contacts.html')


def contact_list(request):
    query = request.GET.get('q', '').strip()
    city = request.GET.get('city', '').strip()
    category = request.GET.get('category', '').strip()
    status = request.GET.get('status', 'active').strip()
    page_number = request.GET.get('page', 1)

    contacts = Contact.objects.all()

    if query:
        from django.db.models import Q
        contacts = contacts.filter(
            Q(business_name__icontains=query) |
            Q(owner_name__icontains=query) |
            Q(phone__icontains=query)
        )
    if city:
        contacts = contacts.filter(city=city)
    if category:
        contacts = contacts.filter(category=category)
    if status == 'active':
        contacts = contacts.filter(is_active=True)
    elif status == 'inactive':
        contacts = contacts.filter(is_active=False)

    # order_by is required before pagination — Postgres gives no row-order
    # guarantee otherwise, which means duplicate/skipped rows across pages.
    contacts = contacts.order_by('business_name')

    total = contacts.count()

    paginator = Paginator(contacts, 25)
    contacts_page = paginator.get_page(page_number)

    # Carry active filters into pagination links so page 2+ doesn't drop them.
    params = request.GET.copy()
    params.pop('page', None)
    filter_string = params.urlencode()

    cities = Contact.objects.values_list('city', flat=True).distinct().exclude(city=None).order_by('city')
    categories = Contact.objects.values_list('category', flat=True).distinct().exclude(category=None).order_by('category')

    return render(request, 'campaigns/contact_list.html', {
        'contacts': contacts_page,
        'cities': cities,
        'categories': categories,
        'query': query,
        'selected_city': city,
        'selected_category': category,
        'selected_status': status,
        'total': total,
        'filter_string': filter_string,
    })


def toggle_contact(request, contact_id):
    if request.method == 'POST':
        contact = get_object_or_404(Contact, id=contact_id)
        contact.is_active = not contact.is_active
        contact.save()
        status = 'activated' if contact.is_active else 'deactivated'
        django_messages.success(request, f'{contact.business_name} {status} successfully.')
    return redirect('contact_list')


def edit_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        business_name = request.POST.get('business_name', '').strip()
        owner_name = request.POST.get('owner_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        city = request.POST.get('city', '').strip()
        category = request.POST.get('category', '').strip()
        email = request.POST.get('email', '').strip()

        if not business_name or not phone:
            django_messages.error(request, 'Business name and phone are required.')
            return render(request, 'campaigns/edit_contact.html', {'contact': contact})

        # Check phone uniqueness — exclude current contact
        if Contact.objects.filter(phone=phone).exclude(id=contact_id).exists():
            django_messages.error(request, f'{phone} is already assigned to another contact.')
            return render(request, 'campaigns/edit_contact.html', {'contact': contact})

        contact.business_name = business_name
        contact.owner_name = owner_name if owner_name else None
        contact.phone = phone
        contact.city = city.title() if city else None
        contact.category = category if category else None
        contact.email = email if email else None
        contact.save()

        django_messages.success(request, f'{contact.business_name} updated successfully.')
        return redirect('contact_list')

    return render(request, 'campaigns/edit_contact.html', {'contact': contact})